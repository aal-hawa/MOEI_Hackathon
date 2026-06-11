"""Excel export (openpyxl): raw verified fields + sources + as-of dates."""
from __future__ import annotations

from io import BytesIO
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_excel(country: str, iso3: str, rows: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dossier"
    headers = ["Domain", "Field", "Value", "Unit", "Source", "As of",
               "Confidence", "Corroborated", "Change", "By"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="9C7A2D")
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([
            r.get("domain"), r.get("field_name"), r.get("value"), r.get("unit"),
            r.get("source_name"), r.get("as_of_date"), r.get("confidence"),
            "yes" if r.get("corroborated") else "no",
            r.get("change_type"), r.get("changed_by"),
        ])
    for col, w in zip("ABCDEFGHIJ", [22, 30, 24, 12, 26, 10, 12, 13, 12, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
